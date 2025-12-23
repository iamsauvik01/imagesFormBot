from openpyxl import load_workbook

wb = load_workbook("data.xlsx")
ws = wb["dataSheet"]

for row in ws.iter_rows():
    cell = row[0]  # Column A
    if isinstance(cell.value, float):
        cell.value = int(cell.value)

wb.save("data.xlsx")
