from openpyxl import Workbook,load_workbook

# wb = Workbook()
# ws = wb.active
# ws.title = "Form Data"

# ws.append(['Sauvik','Das',25,80000,'Software Engineer'])

# wb.save("output.xlsx")


def excel_write(rows):  
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Data"
    
    for data in rows:
        ws.append(data)

    wb.save("output.xlsx")
    