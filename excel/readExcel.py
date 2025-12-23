from openpyxl import load_workbook


def read_excel_rows(path="data.xlsx"):
    wb = load_workbook(path)
    sheet = wb["dataSheet"]

    rows = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rows.append({
            "qr_no": str(row[0]).strip(),
            "name": row[1],
            "designation": row[2],
            "company" : row[3],
            "website" : row[4],
            "address" : row[5],
            "email" : row[6],
            "contact" : row[7]
        })

    # for r in rows:
    #     print(r)

    return rows


if __name__ == "__main__":
    read_excel_rows()
